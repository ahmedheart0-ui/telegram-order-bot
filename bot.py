import asyncio
import os
import re
from datetime import datetime
import openpyxl

from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import Command
from aiogram.fsm.state import StatesGroup, State
from aiogram.fsm.context import FSMContext
from aiogram.utils.keyboard import InlineKeyboardBuilder
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton

# ========= إعدادات =========
BOT_TOKEN = "8418046765:AAFYLds4Bu-1jJPY9BDRJ4fEdGiUORg9978"
GROUP_ID = -1003686759425
STAFF_IDS = {901390292, 7225092840}

EXCEL_FILE = "orders.xlsx"

bot = Bot(BOT_TOKEN)
dp = Dispatcher()

# ========= Excel =========
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([
            "التاريخ","الموظفة","النوع","التفاصيل",
            "العنوان","الأرقام","السعر","message_id"
        ])
        wb.save(EXCEL_FILE)

def save_excel_smart(data, msg_id):
    init_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        data["staff"],
        data["type"],
        ", ".join(data["colors"]),
        data["address"],
        " | ".join(data["phones"]),
        data["price"],
        msg_id
    ])
    wb.save(EXCEL_FILE)

def delete_from_excel(msg_id):
    if not os.path.exists(EXCEL_FILE):
        return
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=8).value == msg_id:
            ws.delete_rows(row)
            break
    wb.save(EXCEL_FILE)

# ========= الواجهة =========
MENU = ReplyKeyboardMarkup(
    keyboard=[[KeyboardButton(text="➕ طلب جديد")]],
    resize_keyboard=True
)

# ========= الحالات (الطلب العادي فقط) =========
class Order(StatesGroup):
    name = State()
    phone = State()
    city = State()
    area = State()
    item = State()
    qty = State()
    price = State()
    photo = State()
    confirm = State()

# ========= start =========
@dp.message(Command("start"))
async def start(message: types.Message):
    if message.from_user.id not in STAFF_IDS:
        return await message.answer("❌ هذا البوت مخصص للموظفات فقط")
    await message.answer("أهلاً 🌸", reply_markup=MENU)

# ========= الطلب العادي (ما تغيّر) =========
@dp.message(lambda m: m.text == "➕ طلب جديد")
async def new_order(message: types.Message, state: FSMContext):
    await state.clear()
    await state.update_data(staff=message.from_user.full_name)
    await state.set_state(Order.name)
    await message.answer("🧑 اسم الزبون:")

@dp.message(Order.name)
async def step_name(m, state):
    await state.update_data(name=m.text)
    await state.set_state(Order.phone)
    await m.answer("📞 رقم الهاتف:")

@dp.message(Order.phone)
async def step_phone(m, state):
    await state.update_data(phone=m.text)
    await state.set_state(Order.city)
    await m.answer("🏙️ المحافظة:")

@dp.message(Order.city)
async def step_city(m, state):
    await state.update_data(city=m.text)
    await state.set_state(Order.area)
    await m.answer("📍 المنطقة:")

@dp.message(Order.area)
async def step_area(m, state):
    await state.update_data(area=m.text)
    await state.set_state(Order.item)
    await m.answer("🛍️ نوع البضاعة:")

@dp.message(Order.item)
async def step_item(m, state):
    await state.update_data(item=m.text)
    await state.set_state(Order.qty)
    await m.answer("🔢 العدد:")

@dp.message(Order.qty)
async def step_qty(m, state):
    await state.update_data(qty=m.text)
    await state.set_state(Order.price)
    await m.answer("💰 السعر:")

@dp.message(Order.price)
async def step_price(m, state):
    await state.update_data(price=m.text)
    await state.set_state(Order.photo)
    await m.answer("📸 أرسلي صورة الطلب:")

@dp.message(Order.photo, F.photo)
async def step_photo(m, state):
    await state.update_data(photo=m.photo[-1].file_id)
    data = await state.get_data()

    text = (
        "📦 طلب جديد – لمسة أنوثة\n"
        "━━━━━━━━━━━━━━\n"
        f"👩‍💼 {data['staff']}\n"
        f"🧑 {data['name']}\n"
        f"📞 {data['phone']}\n"
        f"🏙️ {data['city']} - {data['area']}\n"
        f"🛍️ {data['item']}\n"
        f"🔢 {data['qty']}\n"
        f"💰 {data['price']}\n"
    )

    kb = InlineKeyboardBuilder()
    kb.button(text="✅ نشر الطلب", callback_data="publish")
    kb.button(text="❌ إلغاء الطلب", callback_data="cancel")
    kb.adjust(2)

    await state.set_state(Order.confirm)
    await m.answer_photo(data["photo"], caption=text, reply_markup=kb.as_markup())

@dp.callback_query(Order.confirm)
async def confirm(cb, state):
    data = await state.get_data()

    if cb.data == "publish":
        kb = InlineKeyboardBuilder()
        kb.button(text="❌ إلغاء الطلب", callback_data="delete_after_publish")

        msg = await bot.send_photo(
            GROUP_ID,
            data["photo"],
            caption=cb.message.caption,
            reply_markup=kb.as_markup()
        )

        await cb.message.answer("✅ تم نشر الطلب")

    else:
        await cb.message.delete()

    await state.clear()
    await cb.answer()

# ========= الطلب الذكي (نص حر + صورة) =========
def smart_parse(text):
    data = {
        "type": "",
        "colors": [],
        "address": "",
        "phones": [],
        "price": ""
    }

    lines = [l.strip() for l in text.splitlines() if l.strip()]

    for line in lines:
        if "نوع" in line:
            data["type"] = line.split(":",1)[-1].strip()
        elif "سعر" in line or "الف" in line:
            data["price"] = line
        elif re.search(r"(077|078|079|\+964)", line):
            nums = re.findall(r"(?:\+964|0)\d{9,10}", line)
            data["phones"].extend(nums)
        elif any(x in line for x in ["بغداد","البصرة","النجف","كربلاء","شارع","قرب","حي","منطقة"]):
            data["address"] += line + " "
        elif len(line.split()) <= 3:
            data["colors"].append(line)

    return data

@dp.message(
    lambda m: (
        m.from_user.id in STAFF_IDS
        and (
            (m.text and len(m.text) > 30 and "//" not in m.text)
            or (m.caption and len(m.caption) > 30)
        )
    )
)
async def smart_order(message: types.Message):
    text = message.text or message.caption
    parsed = smart_parse(text)

    order_text = (
        "📦 طلب جديد – لمسة أنوثة\n"
        "━━━━━━━━━━━━━━\n"
        f"👩‍💼 الموظفة: {message.from_user.full_name}\n"
        f"🛍️ النوع: {parsed['type'] or 'غير محدد'}\n"
        f"🎨 التفاصيل: {', '.join(parsed['colors'])}\n"
        f"📍 العنوان: {parsed['address']}\n"
        f"📞 الأرقام:\n" + ("\n".join(parsed['phones']) if parsed['phones'] else "غير مذكور") + "\n"
        f"💰 السعر: {parsed['price'] or 'غير مذكور'}\n"
    )

    kb = InlineKeyboardBuilder()
    kb.button(text="❌ إلغاء الطلب", callback_data="delete_after_publish")
    kb.button(text="🚚 رفع الطلب", callback_data="send_to_delivery")
    kb.adjust(2)

    if message.photo:
        msg = await bot.send_photo(
            GROUP_ID,
            message.photo[-1].file_id,
            caption=order_text,
            reply_markup=kb.as_markup()
        )
    else:
        msg = await bot.send_message(
            GROUP_ID,
            order_text,
            reply_markup=kb.as_markup()
        )

    save_excel_smart(
        {
            "staff": message.from_user.full_name,
            "type": parsed["type"],
            "colors": parsed["colors"],
            "address": parsed["address"],
            "phones": parsed["phones"],
            "price": parsed["price"],
        },
        msg.message_id
    )

    await message.answer("✅ تم نشر الطلب وحفظه")

# ========= إلغاء بعد النشر (للجميع) =========
@dp.callback_query(lambda c: c.data == "delete_after_publish")
async def delete_after_publish(cb: types.CallbackQuery):
    msg_id = cb.message.message_id
    await bot.delete_message(GROUP_ID, msg_id)
    delete_from_excel(msg_id)
   
    await cb.answer("❌ تم إلغاء الطلب من كل مكان", show_alert=True)

    
@dp.callback_query(lambda c: c.data == "send_to_delivery")
async def send_to_delivery(cb: types.CallbackQuery):
    await cb.answer(
        "⏳ سيتم ربط رفع الطلب مع شركة التوصيل لاحقاً",
        show_alert=True
    )

# ========= تشغيل =========
async def main():
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
